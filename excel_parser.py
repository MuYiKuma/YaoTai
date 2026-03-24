from __future__ import annotations

from dataclasses import fields
import re
from typing import Any

from storage_site_input import StorageSiteInput

WorkbookCell = tuple[str, str, int, int, Any]

SHEET_TYPE_TOKENS: dict[str, tuple[str, ...]] = {
    "case_info": ("案場資訊", "基本資訊", "客戶資訊"),
    "schedule": ("24小時", "排程"),
    "strategy": ("24小時", "策略試算"),
    "annual_result": ("年度收益", "年收益"),
}
SHEET_MATCH_MODE: dict[str, str] = {
    "case_info": "any",
    "schedule": "all",
    "strategy": "all",
    "annual_result": "any",
}

LABEL_GROUPS: dict[str, tuple[str, ...]] = {
    "contract_capacity_kw": ("用戶經常契約容量", "契約容量", "經常契約容量"),
    "power_kw": (
        "PCS功率充放電最大功率",
        "儲能系統充放電功率",
        "充放電功率",
        "額定輸出功率",
    ),
    "capacity_kwh": ("電池度數", "儲能額定可儲存能量", "額定可儲存能量", "總額定儲能容量"),
    "dod": ("充放電深度", "DoD", "depth of discharge"),
    "efficiency": ("效率", "轉換效率", "round trip efficiency", "RTE"),
    "summer_spread": ("平日尖峰 離峰", "平日尖峰-離峰", "夏月價差"),
    "non_summer_spread": ("非夏月價差", "平日尖峰 離峰", "平日尖峰-離峰"),
    "dr_capacity_kw": ("抑低契約容量", "需量反應負載管理措施抑低容量", "需量反應容量"),
    "dr_hours": ("參加時數", "需量反應時數", "連續2小時", "連續4小時", "連續6小時"),
    "dr_rate": ("費率", "流動電費扣減費率", "需量反應費率"),
    "dr_execution_rate": ("當日執行率", "需量反應執行率", "執行率"),
    "dr_discount_ratio": ("扣減比率", "需量反應折扣比例"),
    "sr_capacity_kw": ("電力交易平台輔助服務投標容量", "投標容量", "即時備轉容量"),
    "sr_price": ("首年容量費", "容量費"),
    "sr_hours_per_day": ("待命時數", "即時備轉每日時數", "運行小時", "即時備轉時數"),
    "sr_execution_rate": ("投標執行率", "即時備轉執行率"),
}

DEFAULT_SEARCH_OFFSETS: tuple[tuple[int, int], ...] = ((0, 1), (0, 2), (1, 0), (2, 0))

PERCENTAGE_FIELDS = {
    "dod",
    "efficiency",
    "dr_execution_rate",
    "dr_discount_ratio",
    "sr_execution_rate",
}

VALUE_RANGES: dict[str, tuple[float, float]] = {
    "contract_capacity_kw": (100, 50000),
    "power_kw": (10, 10000),
    "capacity_kwh": (10, 50000),
    "dod": (0, 1),
    "efficiency": (0, 1),
    "summer_spread": (0, 20),
    "non_summer_spread": (0, 20),
    "summer_cycles_per_day": (0, 5),
    "non_summer_cycles_per_day": (0, 5),
    "dr_capacity_kw": (0, 10000),
    "dr_hours": (0, 24),
    "dr_rate": (0, 20),
    "dr_execution_rate": (0, 1.2),
    "dr_discount_ratio": (0, 2),
    "sr_capacity_kw": (0, 10000),
    "sr_price": (0, 10000),
    "sr_hours_per_day": (0, 24),
    "sr_execution_rate": (0, 1),
}

_NUMBER_PATTERN = re.compile(r"[-+]?\d+(?:\.\d+)?")


def _load_workbook(file_path: str) -> Any:
    from openpyxl import load_workbook

    return load_workbook(file_path, data_only=True, read_only=True)


def normalize_sheet_name(name: str) -> str:
    """Normalize sheet name by trimming spaces and collapsing whitespace."""
    return re.sub(r"\s+", " ", str(name).strip())


def _normalize_text(value: Any) -> str:
    text = str(value).strip().lower()
    return re.sub(r"[_\-\s]+", " ", text)


def sheet_matches_tokens(sheet_name: str, tokens: list[str] | tuple[str, ...]) -> bool:
    """Return True if all tokens appear in normalized sheet name."""
    normalized_sheet = _normalize_text(normalize_sheet_name(sheet_name))
    normalized_tokens = [_normalize_text(token) for token in tokens]
    if any(token in ("策略試算", "排程") for token in tokens):
        return all(token in normalized_sheet for token in normalized_tokens)
    return any(token in normalized_sheet for token in normalized_tokens)


def find_sheets_by_type(workbook: Any) -> dict[str, list[Any]]:
    """Classify workbook sheets into predefined sheet types."""
    result: dict[str, list[Any]] = {name: [] for name in SHEET_TYPE_TOKENS}
    for sheet in workbook.worksheets:
        normalized_name = normalize_sheet_name(sheet.title)
        print(f"[excel_parser] 掃描 sheet: {normalized_name}")

        for sheet_type, tokens in SHEET_TYPE_TOKENS.items():
            match_mode = SHEET_MATCH_MODE[sheet_type]
            normalized_tokens = [_normalize_text(token) for token in tokens]
            normalized_sheet = _normalize_text(normalized_name)
            if match_mode == "all":
                matched = all(token in normalized_sheet for token in normalized_tokens)
            else:
                matched = any(token in normalized_sheet for token in normalized_tokens)
            if matched:
                result[sheet_type].append(sheet)
                print(f"[excel_parser] 辨識為 {sheet_type}: {normalized_name}")

    return result


def _extract_numeric(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    match = _NUMBER_PATTERN.search(str(value).replace(",", ""))
    if match is None:
        return None
    return float(match.group())


def load_workbook_data(file_path: str) -> list[WorkbookCell]:
    """Return all non-empty cells in workbook."""
    workbook = _load_workbook(file_path)
    results: list[WorkbookCell] = []

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and not cell.value.strip():
                    continue
                results.append((sheet.title, cell.coordinate, cell.row, cell.column, cell.value))

    return results


def find_label_cell(sheet: Any, label_candidates: list[str] | tuple[str, ...]) -> Any | None:
    """Find cell matching any candidate label text."""
    normalized_candidates = [_normalize_text(candidate) for candidate in label_candidates]
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.strip():
                continue
            normalized_cell = _normalize_text(cell.value)
            if any(candidate in normalized_cell for candidate in normalized_candidates):
                return cell
    return None


def find_neighbor_numeric_value(
    sheet: Any,
    row: int,
    col: int,
    search_offsets: list[tuple[int, int]] | tuple[tuple[int, int], ...] | None = None,
) -> tuple[float, str | None] | None:
    """Find numeric value around a label cell and return number with source unit text."""
    offsets = search_offsets or DEFAULT_SEARCH_OFFSETS
    for row_offset, col_offset in offsets:
        nearby_cell = sheet.cell(row=row + row_offset, column=col + col_offset)
        numeric_value = _extract_numeric(nearby_cell.value)
        if numeric_value is not None:
            raw_text = nearby_cell.value if isinstance(nearby_cell.value, str) else None
            return numeric_value, raw_text
    return None


def convert_unit_if_needed(field_name: str, value: float, unit_text: str | None) -> float:
    """Convert extracted value according to field/unit rules."""
    normalized_unit = _normalize_text(unit_text or "")

    if field_name == "power_kw":
        if "mw" in normalized_unit:
            return value * 1000
        return value

    if field_name == "capacity_kwh":
        if "mwh" in normalized_unit:
            return value * 1000
        return value

    if field_name in PERCENTAGE_FIELDS and 1 < value <= 100:
        if field_name in {"dr_execution_rate", "sr_execution_rate"}:
            is_percent_text = "%" in (unit_text or "") or "百分" in (unit_text or "")
            if not is_percent_text and value <= 10:
                return value
        return value / 100

    if field_name == "dr_discount_ratio" and value > 100:
        return value / 100

    return value


def validate_parsed_value(field_name: str, value: float) -> tuple[bool, str | None]:
    """Validate parsed value against field range."""
    min_value, max_value = VALUE_RANGES[field_name]
    if min_value <= value <= max_value:
        return True, None
    return False, f"欄位 {field_name} 值 {value} 超出合理範圍 [{min_value}, {max_value}]"


def _parse_spreads_from_strategy(sheet: Any) -> dict[str, tuple[float, str | None]]:
    spread_values: dict[str, tuple[float, str | None]] = {}
    spread_cell = find_label_cell(sheet, LABEL_GROUPS["summer_spread"])
    if spread_cell is None:
        return spread_values

    numeric_values: list[tuple[float, str | None]] = []
    for col_offset in range(1, 6):
        neighbor = sheet.cell(row=spread_cell.row, column=spread_cell.column + col_offset)
        parsed = _extract_numeric(neighbor.value)
        if parsed is None:
            continue
        raw_text = neighbor.value if isinstance(neighbor.value, str) else None
        numeric_values.append((parsed, raw_text))

    if len(numeric_values) >= 1:
        spread_values["summer_spread"] = numeric_values[0]
    if len(numeric_values) >= 2:
        spread_values["non_summer_spread"] = numeric_values[1]

    return spread_values


def _choose_strategy_sheet(strategy_sheets: list[Any], strategy_sheet_name: str | None) -> Any | None:
    if not strategy_sheets:
        return None
    if strategy_sheet_name:
        target = normalize_sheet_name(strategy_sheet_name)
        for sheet in strategy_sheets:
            if normalize_sheet_name(sheet.title) == target:
                return sheet
    for sheet in strategy_sheets:
        if "商二" in normalize_sheet_name(sheet.title):
            return sheet
    return strategy_sheets[0]


def _parse_field_from_sheet(sheet: Any, field_name: str) -> tuple[float, str | None] | None:
    label_cell = find_label_cell(sheet, LABEL_GROUPS[field_name])
    if label_cell is None:
        return None
    return find_neighbor_numeric_value(sheet, label_cell.row, label_cell.column)


def parse_to_storage_input(
    file_path: str,
    strategy_sheet: str | None = None,
) -> tuple[StorageSiteInput, list[str]]:
    """Parse financial model Excel to StorageSiteInput and warnings."""
    workbook = _load_workbook(file_path)
    sheet_map = find_sheets_by_type(workbook)

    case_sheet = sheet_map["case_info"][0] if sheet_map["case_info"] else None
    schedule_sheet = sheet_map["schedule"][0] if sheet_map["schedule"] else None
    selected_strategy_sheet = _choose_strategy_sheet(sheet_map["strategy"], strategy_sheet)

    parsed_values: dict[str, float] = {}
    warnings: list[str] = []

    field_order = [
        "contract_capacity_kw",
        "power_kw",
        "capacity_kwh",
        "dod",
        "efficiency",
        "summer_spread",
        "non_summer_spread",
        "dr_capacity_kw",
        "dr_hours",
        "dr_rate",
        "dr_execution_rate",
        "dr_discount_ratio",
        "sr_capacity_kw",
        "sr_price",
        "sr_hours_per_day",
        "sr_execution_rate",
    ]

    source_sheets = [sheet for sheet in (case_sheet, schedule_sheet, selected_strategy_sheet) if sheet is not None]

    if selected_strategy_sheet is not None:
        spread_pairs = _parse_spreads_from_strategy(selected_strategy_sheet)
        if "summer_spread" in spread_pairs:
            raw_value, unit_text = spread_pairs["summer_spread"]
            converted = convert_unit_if_needed("summer_spread", raw_value, unit_text)
            valid, reason = validate_parsed_value("summer_spread", converted)
            if valid:
                parsed_values["summer_spread"] = converted
                print(f"[excel_parser] 成功抓取 summer_spread={converted}")
            elif reason:
                warnings.append(reason)
        else:
            warnings.append("無法從 strategy sheet 的平日尖峰-離峰列解析 summer_spread。")

        if "non_summer_spread" in spread_pairs:
            raw_value, unit_text = spread_pairs["non_summer_spread"]
            converted = convert_unit_if_needed("non_summer_spread", raw_value, unit_text)
            valid, reason = validate_parsed_value("non_summer_spread", converted)
            if valid:
                parsed_values["non_summer_spread"] = converted
                print(f"[excel_parser] 成功抓取 non_summer_spread={converted}")
            elif reason:
                warnings.append(reason)
        else:
            warnings.append("strategy sheet 找到價差標籤，但缺少非夏月價差欄位。")

    for field_name in field_order:
        if field_name in parsed_values:
            continue

        parsed_result: tuple[float, str | None] | None = None
        for sheet in source_sheets:
            parsed_result = _parse_field_from_sheet(sheet, field_name)
            if parsed_result is not None:
                break

        if parsed_result is None:
            warnings.append(f"找不到欄位 {field_name} 的值，已保留預設值。")
            continue

        raw_value, unit_text = parsed_result
        converted_value = convert_unit_if_needed(field_name, raw_value, unit_text)
        is_valid, reason = validate_parsed_value(field_name, converted_value)
        if not is_valid:
            warnings.append(reason or f"欄位 {field_name} 值 {converted_value} 不合理")
            continue

        parsed_values[field_name] = converted_value
        print(f"[excel_parser] 成功抓取 {field_name}={converted_value}")

        if unit_text is None:
            warnings.append(f"欄位 {field_name} 未明確提供單位，已直接採用數值 {raw_value}。")

    for cycles_field in ("summer_cycles_per_day", "non_summer_cycles_per_day"):
        warnings.append(f"找不到欄位 {cycles_field} 的值，已保留預設值。")
        print(f"[excel_parser] 欄位 {cycles_field} fallback 到預設值")

    init_values: dict[str, float] = {}
    for field in fields(StorageSiteInput):
        if field.name in parsed_values:
            init_values[field.name] = parsed_values[field.name]
        elif field.name in VALUE_RANGES:
            print(f"[excel_parser] 欄位 {field.name} fallback 到預設值")

    return StorageSiteInput(**init_values), warnings
